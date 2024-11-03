#!/usr/bin/env python3
"""Module to update a portfolio tracking sheet."""

from datetime import date
import sys
import requests
import numbers_parser
import openpyxl

from csu_config import CSUConfig
from csu_types import Config, SheetType, ExcelDoc, NumbersDoc, Coin
from csu_helpers import round_float

def main():
    """Main function that contains the update workflow for the tracking sheet."""
    config = Config(CSUConfig())

    if config.sheet_type == SheetType.NUMBERS:
        numbers_doc = load_numbers_doc(config)
        token_list = prepare_dataset(numbers_doc, config)

        coins = []
        for token_set in token_list:
            coins += get_data_from_cmc_api(token_set, config)

        update_numbers_sheet(numbers_doc, coins, config)
    else:
        excel_doc = load_excel_doc(config)
        token_list = prepare_dataset(excel_doc, config)

        coins = []
        for token_set in token_list:
            coins += get_data_from_cmc_api(token_set, config)

        update_excel_sheet(excel_doc, coins, config)

def load_numbers_doc(config: Config) -> NumbersDoc:
    """Returns a csu_types.NumbersDoc from an input file."""
    try:
        # Test path
        with open(config.input_path, encoding="utf-8") as f:
            f.close()
    except OSError:
        sys.exit(f"Failed to open/read input_path file : '{config.input_path}'. "\
                 "Check its presence in the current folder and its contents.")

    doc = numbers_parser.Document(config.input_path)
    sheets = doc.sheets
    table = sheets[config.sheet_index].tables[config.table_name]
    rows = table.rows()

    return NumbersDoc(doc, sheets, table, rows)

def load_excel_doc(config: Config) -> ExcelDoc:
    """Returns a csu_types.ExcelDoc from an input file."""
    try:
        # Test path
        with open(config.input_path, encoding="utf-8") as f:
            f.close()
    except OSError:
        sys.exit(f"Failed to open/read input_path file : '{config.input_path}'. "\
                 "Check its presence in the current folder and its contents.")

    doc = openpyxl.load_workbook(config.input_path)

    sheets = []
    for sheet in doc.worksheets:
        sheets.append(sheet)

    sheet = sheets[config.sheet_index]
    table = sheet.tables[config.table_name]
    range_str = table.ref

    if range_str == "":
        sys.exit("Failed to find table range in the Excel file. Check that values are in a table.")

    range_cells = sheets[config.sheet_index][range_str]
    rows = [[cell.value for cell in row] for row in range_cells]

    return ExcelDoc(doc, sheets, table, rows)

def prepare_dataset(doc, config: Config):
    """Returns all the token names contained in a table as a list of string : ["BTC,ETH,...","..."].
    Required because CMC API allows 100 tokens per HTTPS request."""
    token_list = []
    token_set = []
    cmc_api_limit = 100

    si = config.table_start_row_index
    ei = config.table_end_row_index

    rows = doc.rows

    if ei >= 0:
        # Iterate until last index.
        ei = len(rows)
    elif ei < 0:
        # Remove last(s) index(s).
        ei = len(rows) + ei

    val = ""
    for row in rows[si:ei]:
        if isinstance(doc, NumbersDoc):
            val = row[config.table_coin_name_col_index].value
        else:
            val = row[config.table_coin_name_col_index]

        if len(token_set) < cmc_api_limit:
            token_set.append(val)
        else:
            token_str = ','.join(token_set)
            token_list.append(token_str)
            token_set.clear()
            token_set.append(val)

        si += 1

    token_str = ','.join(token_set)
    token_list.append(token_str)
    return token_list

def get_data_from_cmc_api(token_set: str, config: Config) -> list[Coin]:
    """Gets current price for a list of tokens from CoinMarketCap API."""
    # Get data.
    url = f"{config.cmc_api_url}?symbol={token_set}"
    headers = {"X-CMC_PRO_API_KEY": config.cmc_api_token}
    try:
        response = requests.get(url, headers=headers, timeout=10)
    except (requests.ConnectTimeout, requests.HTTPError, requests.ReadTimeout,\
            requests.Timeout, requests.ConnectionError, requests.exceptions.MissingSchema) as e:
        sys.exit(f"ERROR : str{e}. Check the configured URL.")
    rsp_json = response.json()

    # Return in case of error.
    if "status" not in rsp_json:
        sys.exit(f"ERROR : unable to retrieve 'status' key in response for '{token_set}'. Brut response : {rsp_json}")

    if "error_code" not in rsp_json["status"]:
        sys.exit(f"ERROR : unable to retrieve 'status' key in response for '{token_set}'. Brut response : {rsp_json}")

    if rsp_json["status"]["error_code"] != 0:
        error_message = rsp_json.get("status", {}).get("error_message", "Unknown error")
        sys.exit(f"ERROR : unable to fetch data for '{token_set}': {error_message}")

    # Otherwise formats and returns a list of Coins.
    coins = []
    for coin_name, value in rsp_json["data"].items():
        if len(value) > 0:
            price = value[0]["quote"]["USD"]["price"]
            if price is None:
                print(f"ERROR : no price for {coin_name}, price set to 0. Continuing.")
            rounded_price = round_float(price)
            coins.append(Coin(coin_name, rounded_price))
        else:
            print(f"ERROR : no value for {coin_name}, value set to 0. Continuing.")
            coins.append(Coin(coin_name, 0))

    return coins

def update_numbers_sheet(numbers_doc: NumbersDoc, coins: list[Coin], config: Config):
    """Update the output sheet from fresh data from CMC API."""
    cur_row_index = config.table_start_row_index
    cur_date = date.today().strftime('%d/%m/%Y')

    for coin in coins:
        numbers_doc.table.write(cur_row_index, config.table_coin_price_col_index, coin.price)

        if config.table_date_col_index > -1:
            numbers_doc.table.write(cur_row_index, config.table_date_col_index, cur_date)

        cur_row_index += 1

    numbers_doc.doc.save(config.output_path)

def update_excel_sheet(excel_doc: ExcelDoc, coins: list[Coin], config: Config):
    """Update the output sheet from fresh data from CMC API."""
    cur_row_index = config.table_start_row_index
    cur_date = date.today().strftime('%d/%m/%Y')

    for coin in coins:
        # Excel rows and cols index start at 1
        excel_doc.sheets[0].cell(row=cur_row_index + 1, column=config.table_coin_price_col_index + 1).value = coin.price

        if config.table_date_col_index > -1:
            excel_doc.sheets[0].cell(row=cur_row_index + 1, column=config.table_date_col_index + 1).value = cur_date

        cur_row_index += 1

    excel_doc.doc.save(config.output_path)

if __name__ == "__main__":
    main()
